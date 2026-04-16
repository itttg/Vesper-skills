#!/usr/bin/env python3
"""Update InvoiceList.xlsx from XML files stored in the Files folder.

This script is designed for Vietnamese e-invoice XML files with schema variations.
It extracts review-ready invoice rows, upgrades worksheet headers when needed,
and upserts rows into Invoice_Tax_Lines.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install it with: pip install openpyxl"
    ) from exc

HEADERS = [
    "InvoiceDate",
    "InvoiceNo",
    "SellerCompany",
    "SellerTaxCode",
    "BuyerName",
    "BuyerTaxCode",
    "Description",
    "TaxRate",
    "AmountBeforeTax",
    "TaxAmount",
    "AmountAfterTax",
    "DetailsText",
    "Status",
    "Notes",
    "InvoiceFormNo",
    "InvoiceSeries",
    "InvoiceNumber",
    "TaxAuthorityCode",
    "SigningDate",
    "InvoiceType",
    "InvoiceNature",
    "PaymentMethod",
    "SellerAddress",
    "BuyerAddress",
    "Currency",
    "ExchangeRate",
    "AmountInWords",
    "SourceFile",
]

SHEET_NAME = "Invoice_Tax_Lines"
SUMMARY_FILE = "invoice_update_summary.json"
ERROR_FILE = "invoice_parse_errors.csv"
ZERO = Decimal("0")


@dataclass
class InvoiceRow:
    invoice_date: str = ""
    invoice_no: str = ""
    seller_company: str = ""
    seller_tax_code: str = ""
    buyer_name: str = ""
    buyer_tax_code: str = ""
    description: str = ""
    tax_rate: str = ""
    amount_before_tax: str = ""
    tax_amount: str = ""
    amount_after_tax: str = ""
    details_text: str = ""
    status: str = "OK"
    notes: str = ""
    invoice_form_no: str = ""
    invoice_series: str = ""
    invoice_number: str = ""
    tax_authority_code: str = ""
    signing_date: str = ""
    invoice_type: str = ""
    invoice_nature: str = ""
    payment_method: str = ""
    seller_address: str = ""
    buyer_address: str = ""
    currency: str = ""
    exchange_rate: str = ""
    amount_in_words: str = ""
    source_file: str = ""

    def to_excel_row(self) -> List[str]:
        return [
            self.invoice_date,
            self.invoice_no,
            self.seller_company,
            self.seller_tax_code,
            self.buyer_name,
            self.buyer_tax_code,
            self.description,
            self.tax_rate,
            self.amount_before_tax,
            self.tax_amount,
            self.amount_after_tax,
            self.details_text,
            self.status,
            self.notes,
            self.invoice_form_no,
            self.invoice_series,
            self.invoice_number,
            self.tax_authority_code,
            self.signing_date,
            self.invoice_type,
            self.invoice_nature,
            self.payment_method,
            self.seller_address,
            self.buyer_address,
            self.currency,
            self.exchange_rate,
            self.amount_in_words,
            self.source_file,
        ]

    def dedupe_key(self) -> Tuple[str, str, str, str]:
        return (
            normalize_compare(self.invoice_no),
            normalize_compare(self.description),
            normalize_compare(self.tax_rate),
            normalize_compare(self.amount_before_tax),
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse XML invoices from Files and update InvoiceList.xlsx"
    )
    parser.add_argument("folder", help="Working folder containing Files and InvoiceList.xlsx")
    parser.add_argument("--files-dir", default="Files")
    parser.add_argument("--workbook", default="InvoiceList.xlsx")
    parser.add_argument("--sheet-name", default=SHEET_NAME)
    parser.add_argument("--disable-dedupe", action="store_true")
    parser.add_argument("--details-max-len", type=int, default=30000)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.folder).expanduser().resolve()
    files_dir = base_dir / args.files_dir
    workbook_path = base_dir / args.workbook

    validate_inputs(base_dir, files_dir, workbook_path)

    xml_files = sorted(files_dir.glob("*.xml"))
    summary: Dict[str, object] = {
        "runAt": datetime.now().isoformat(timespec="seconds"),
        "baseDir": str(base_dir),
        "filesDir": str(files_dir),
        "workbook": str(workbook_path),
        "sheetName": args.sheet_name,
        "xmlFilesFound": len(xml_files),
        "xmlParsedSuccess": 0,
        "rowsPrepared": 0,
        "rowsWritten": 0,
        "rowsUpdated": 0,
        "duplicatesSkipped": 0,
        "parseErrors": 0,
        "errorFiles": [],
    }

    parsed_rows: List[InvoiceRow] = []
    error_rows: List[Dict[str, str]] = []

    for xml_path in xml_files:
        try:
            rows = parse_invoice_file(xml_path, args.details_max_len)
            if not rows:
                raise ValueError("No invoice rows produced from XML")
            parsed_rows.extend(rows)
            summary["xmlParsedSuccess"] = int(summary["xmlParsedSuccess"]) + 1
        except Exception as exc:  # pragma: no cover
            error_rows.append({"FileName": xml_path.name, "Error": str(exc)})
            summary["parseErrors"] = int(summary["parseErrors"]) + 1
            cast_list(summary["errorFiles"]).append(xml_path.name)

    summary["rowsPrepared"] = len(parsed_rows)

    written_count, updated_count, duplicate_count = upsert_workbook(
        workbook_path=workbook_path,
        sheet_name=args.sheet_name,
        rows=parsed_rows,
        disable_dedupe=args.disable_dedupe,
    )
    summary["rowsWritten"] = written_count
    summary["rowsUpdated"] = updated_count
    summary["duplicatesSkipped"] = duplicate_count

    write_summary(base_dir / SUMMARY_FILE, summary)
    if error_rows:
        write_error_csv(base_dir / ERROR_FILE, error_rows)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


def validate_inputs(base_dir: Path, files_dir: Path, workbook_path: Path) -> None:
    if not base_dir.exists():
        raise SystemExit(f"Folder does not exist: {base_dir}")
    if not files_dir.exists():
        raise SystemExit(f"Files folder does not exist: {files_dir}")
    if not workbook_path.exists():
        create_workbook(workbook_path)


def create_workbook(workbook_path: Path) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(workbook_path)


def upsert_workbook(
    workbook_path: Path,
    sheet_name: str,
    rows: Sequence[InvoiceRow],
    disable_dedupe: bool,
) -> Tuple[int, int, int]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    upgrade_headers(ws)
    header_index = header_index_map(ws)
    existing = {} if disable_dedupe else load_existing_rows(ws, header_index)

    written = 0
    updated = 0
    duplicates = 0

    for row in rows:
        key = row.dedupe_key()
        if disable_dedupe or key not in existing:
            ws.append(row.to_excel_row())
            written += 1
            if not disable_dedupe:
                existing[key] = ws.max_row
            continue

        duplicates += 1

    wb.save(workbook_path)
    return written, updated, duplicates


def upgrade_headers(ws) -> None:
    if ws.max_row == 0:
        ws.append(HEADERS)
        return

    existing_values = [ws.cell(1, idx).value for idx in range(1, ws.max_column + 1)]
    if not any(existing_values):
        for idx, value in enumerate(HEADERS, start=1):
            ws.cell(1, idx).value = value
        return

    existing_text = [str(v).strip() if v is not None else "" for v in existing_values]
    for idx, header in enumerate(HEADERS, start=1):
        current = existing_text[idx - 1] if idx - 1 < len(existing_text) else ""
        if not current:
            ws.cell(1, idx).value = header


def header_index_map(ws) -> Dict[str, int]:
    return {
        str(ws.cell(1, idx).value).strip(): idx
        for idx in range(1, ws.max_column + 1)
        if ws.cell(1, idx).value is not None
    }


def load_existing_rows(ws, header_index: Dict[str, int]) -> Dict[Tuple[str, str, str, str], int]:
    required = ["InvoiceNo", "Description", "TaxRate", "AmountBeforeTax"]
    if not all(name in header_index for name in required):
        return {}

    result: Dict[Tuple[str, str, str, str], int] = {}
    for row_idx in range(2, ws.max_row + 1):
        key = (
            normalize_compare(str(ws.cell(row_idx, header_index["InvoiceNo"]).value or "")),
            normalize_compare(str(ws.cell(row_idx, header_index["Description"]).value or "")),
            normalize_compare(str(ws.cell(row_idx, header_index["TaxRate"]).value or "")),
            normalize_compare(str(ws.cell(row_idx, header_index["AmountBeforeTax"]).value or "")),
        )
        if any(key):
            result[key] = row_idx
    return result


def update_existing_row(ws, row_idx: int, header_index: Dict[str, int], new_row: InvoiceRow) -> bool:
    changed = False
    values = dict(zip(HEADERS, new_row.to_excel_row()))
    for header, value in values.items():
        col_idx = header_index.get(header)
        if not col_idx:
            continue
        existing = ws.cell(row_idx, col_idx).value
        if should_fill_existing(existing, value):
            ws.cell(row_idx, col_idx).value = value
            changed = True
    return changed


def should_fill_existing(existing: object, new_value: str) -> bool:
    existing_text = clean_space(str(existing)) if existing is not None else ""
    new_text = clean_space(new_value)
    if not new_text:
        return False
    if not existing_text:
        return True
    return existing_text != new_text and len(existing_text) < len(new_text)


def parse_invoice_file(xml_path: Path, details_max_len: int) -> List[InvoiceRow]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    strip_namespaces(root)

    details_text = extract_details_text(root, details_max_len)
    invoice_date = first_non_empty(
        format_date(find_text(root, ["NLap", "NgayLap", "InvoiceDate"])),
        extract_by_regex(details_text, r"Ngày\s+(\d{1,2}[/-]\d{1,2}[/-]\d{4})"),
    )
    signing_date = first_non_empty(
        format_date(find_text(root, ["NKy", "SigningTime", "NgayKy"])),
        extract_signing_date_from_signature(details_text),
    )

    invoice_form_no = safe_invoice_number(find_text(root, ["KHMSHDon", "KHMShDon", "InvoiceFormNo"]))
    invoice_series = first_non_empty(
        find_text(root, ["KHHDon", "KyHieu", "InvoiceSeries"]),
        extract_series_from_filename(xml_path.stem),
    )
    invoice_number = safe_invoice_number(
        first_non_empty(
            find_text(root, ["SHDon", "SoHDon", "InvoiceNumber", "SoHoaDon"]),
            extract_number_from_filename(xml_path.stem),
        )
    )
    invoice_no = build_invoice_no(invoice_series, invoice_number, xml_path.stem)

    tax_authority_code = find_text(root, ["MCCQT", "MCQT", "TaxAuthorityCode"])
    invoice_type = first_non_empty(find_text(root, ["THDon", "TLHDon", "InvoiceType"]))
    invoice_nature = map_invoice_nature(find_text(root, ["TChat", "InvoiceNature"]))
    payment_method = find_text(root, ["HTTToan", "PaymentMethod"])
    currency = find_text(root, ["DVTTe", "Currency"])
    exchange_rate = format_decimal_plain(parse_decimal(find_text(root, ["TGia", "ExchangeRate"])))
    amount_in_words = find_text(root, ["TgTTTBChu", "TotalAmountInWords"])

    seller_company, seller_tax_code, seller_address = extract_party(root, role="seller", details_text=details_text)
    buyer_name, buyer_tax_code, buyer_address = extract_party(root, role="buyer", details_text=details_text)

    total_before_tax = parse_decimal(find_text(root, ["TgTCThue", "TotalAmountWithoutTax"]))
    total_tax = parse_decimal(find_text(root, ["TgTThue", "TotalVATAmount", "TaxAmount"]))
    total_after_tax = parse_decimal(find_text(root, ["TgTTTBSo", "TotalAmountWithTax", "TotalPaymentAmount"]))
    if total_after_tax == ZERO and total_before_tax != ZERO:
        total_after_tax = total_before_tax + total_tax

    tax_summary_lines = extract_tax_summary_lines(root)
    line_items = extract_line_items(root)

    common = {
        "invoice_date": invoice_date,
        "invoice_no": invoice_no,
        "seller_company": seller_company,
        "seller_tax_code": seller_tax_code,
        "buyer_name": buyer_name,
        "buyer_tax_code": buyer_tax_code,
        "details_text": details_text,
        "status": "OK" if invoice_no else "REVIEW",
        "invoice_form_no": invoice_form_no,
        "invoice_series": invoice_series,
        "invoice_number": invoice_number,
        "tax_authority_code": tax_authority_code,
        "signing_date": signing_date,
        "invoice_type": invoice_type,
        "invoice_nature": invoice_nature,
        "payment_method": payment_method,
        "seller_address": seller_address,
        "buyer_address": buyer_address,
        "currency": currency,
        "exchange_rate": exchange_rate,
        "amount_in_words": amount_in_words,
        "source_file": xml_path.name,
    }

    grouped_source = tax_summary_lines if tax_summary_lines else line_items
    if grouped_source:
        grouped: Dict[str, Dict[str, object]] = defaultdict(
            lambda: {"description": [], "before_tax": ZERO, "tax": ZERO, "after_tax": ZERO}
        )
        for item in grouped_source:
            rate = str(item.get("tax_rate") or "N/A")
            desc = str(item.get("description") or "")
            amount = decimal_or_zero(item.get("amount_before_tax"))
            tax_amount = decimal_or_zero(item.get("tax_amount"))
            after_tax = decimal_or_zero(item.get("amount_after_tax"))
            bucket = grouped[rate]
            if desc:
                cast_list(bucket["description"]).append(desc)
            bucket["before_tax"] = decimal_or_zero(bucket["before_tax"]) + amount
            bucket["tax"] = decimal_or_zero(bucket["tax"]) + tax_amount
            bucket["after_tax"] = decimal_or_zero(bucket["after_tax"]) + after_tax

        rows: List[InvoiceRow] = []
        for rate, bucket in grouped.items():
            before_tax = decimal_or_zero(bucket["before_tax"])
            tax_amount = decimal_or_zero(bucket["tax"])
            if tax_amount == ZERO:
                numeric_rate = parse_tax_rate(rate)
                if numeric_rate is not None:
                    tax_amount = quantize_money(before_tax * numeric_rate / Decimal("100"))
            after_tax = decimal_or_zero(bucket["after_tax"])
            if after_tax == ZERO:
                after_tax = before_tax + tax_amount
            descriptions = unique_preserve_order(cast_list(bucket["description"]))
            rows.append(
                InvoiceRow(
                    **common,
                    description="; ".join(descriptions) or "Summary",
                    tax_rate=rate,
                    amount_before_tax=format_money(before_tax),
                    tax_amount=format_money(tax_amount),
                    amount_after_tax=format_money(after_tax),
                    notes=f"SourceXML={xml_path.name}; ParseMode=vat-rate-grouped",
                )
            )
        return rows

    return [
        InvoiceRow(
            **common,
            description="Summary",
            tax_rate=extract_tax_rate_from_text(details_text) or "N/A",
            amount_before_tax=format_money(total_before_tax),
            tax_amount=format_money(total_tax),
            amount_after_tax=format_money(total_after_tax),
            notes=f"SourceXML={xml_path.name}; ParseMode=summary-fallback",
        )
    ]


def extract_tax_summary_lines(root: ET.Element) -> List[Dict[str, object]]:
    items: List[Dict[str, object]] = []
    candidate_names = {"LTSuat", "LTSuat", "TTLTSuat", "ThTTLTSuat"}
    for element in root.iter():
        if element.tag not in candidate_names:
            continue
        tax_rate = normalize_tax_rate(child_text(element, ["TSuat", "ThueSuat", "TaxRate", "VATRate"]))
        amount_before_tax = parse_decimal(child_text(element, ["ThTien", "THTien", "ThanhTien", "Amount", "TotalAmount"]))
        tax_amount = parse_decimal(child_text(element, ["TThue", "TaxAmount", "VATAmount"]))
        after_tax = amount_before_tax + tax_amount if amount_before_tax or tax_amount else ZERO
        if not tax_rate and amount_before_tax == ZERO and tax_amount == ZERO:
            continue
        items.append(
            {
                "description": "Summary",
                "tax_rate": tax_rate or "N/A",
                "amount_before_tax": amount_before_tax,
                "tax_amount": tax_amount,
                "amount_after_tax": after_tax,
            }
        )
    return items


def extract_line_items(root: ET.Element) -> List[Dict[str, object]]:
    items: List[Dict[str, object]] = []
    candidate_names = {"HHDVu", "HHDon", "Item", "Row", "CTietHDon"}
    for element in root.iter():
        if element.tag not in candidate_names:
            continue
        description = first_non_empty(
            child_text(element, ["THHDVu", "Ten", "TenHHDVu", "Description", "DVTTe", "TSuat"]),
            text_join(element),
        )
        if not description:
            continue
        amount_before_tax = parse_decimal(child_text(element, ["ThTien", "Amount", "TotalAmount", "ThanhTien", "THTien"]))
        tax_rate = normalize_tax_rate(child_text(element, ["TSuat", "ThueSuat", "TaxRate", "VATRate"]))
        tax_amount = parse_decimal(child_text(element, ["TThue", "TaxAmount", "VATAmount"]))
        after_tax = amount_before_tax + tax_amount if amount_before_tax or tax_amount else ZERO
        if amount_before_tax == ZERO and tax_amount == ZERO and len(clean_space(description)) < 6:
            continue
        items.append(
            {
                "description": shorten(clean_space(description), 500),
                "tax_rate": tax_rate or "N/A",
                "amount_before_tax": amount_before_tax,
                "tax_amount": tax_amount,
                "amount_after_tax": after_tax,
            }
        )
    return items


def extract_party(root: ET.Element, role: str, details_text: str) -> Tuple[str, str, str]:
    if role == "seller":
        tags = ["NBan", "Seller", "NguoiBan"]
        name_patterns = [r"Tên đơn vị bán hàng\s*:?\s*(.+)"]
    else:
        tags = ["NMua", "Buyer", "NguoiMua"]
        name_patterns = [r"Tên đơn vị mua hàng\s*:?\s*(.+)", r"Tên đơn vị\s*:?\s*(.+)"]

    for element in root.iter():
        if element.tag in tags:
            name = first_non_empty(child_text(element, ["Ten", "CompanyName", "TenDV", "Name"]))
            tax_code = first_non_empty(child_text(element, ["MST", "TaxCode", "MaSoThue"]))
            address = first_non_empty(child_text(element, ["DChi", "Address", "DiaChi"]))
            if name or tax_code or address:
                return clean_space(name), clean_tax_code(tax_code), clean_space(address)

    name = ""
    for pattern in name_patterns:
        match = re.search(pattern, details_text, flags=re.IGNORECASE)
        if match:
            name = clean_space(match.group(1))
            break
    tax_matches = re.findall(r"Mã số thuế\s*:?\s*(\d{10,14})", details_text, flags=re.IGNORECASE)
    tax_code = clean_tax_code(tax_matches[0 if role == "seller" else min(1, len(tax_matches) - 1)]) if tax_matches else ""
    return name, tax_code, ""


def strip_namespaces(root: ET.Element) -> None:
    for element in root.iter():
        if "}" in element.tag:
            element.tag = element.tag.split("}", 1)[1]


def find_text(root: ET.Element, candidate_tags: Sequence[str]) -> str:
    lowered = {tag.lower() for tag in candidate_tags}
    for element in root.iter():
        if element.tag.lower() in lowered:
            value = text_join(element)
            if value:
                return value
    return ""


def child_text(element: ET.Element, candidate_tags: Sequence[str]) -> str:
    lowered = {tag.lower() for tag in candidate_tags}
    for child in element.iter():
        if child is element:
            continue
        if child.tag.lower() in lowered:
            value = text_join(child)
            if value:
                return value
    return ""


def extract_details_text(root: ET.Element, max_len: int) -> str:
    text = clean_space("\n".join(text for text in root.itertext() if clean_space(text)))
    return shorten(text, max_len)


def text_join(element: ET.Element) -> str:
    return clean_space(" ".join(text.strip() for text in element.itertext() if text and text.strip()))


def clean_space(value: Optional[str]) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def shorten(value: str, max_len: int) -> str:
    if len(value) <= max_len:
        return value
    return value[: max_len - 3].rstrip() + "..."


def extract_series_from_filename(stem: str) -> str:
    match = re.match(r"([A-Za-z0-9]+)_", stem)
    return match.group(1) if match else ""


def extract_number_from_filename(stem: str) -> str:
    match = re.search(r"_([0-9]+)$", stem)
    return match.group(1) if match else ""


def safe_invoice_number(value: str) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", value)
    return digits or clean_space(value)


def build_invoice_no(series: str, number: str, stem: str) -> str:
    if series and number:
        return f"{series}_{number}"
    if series:
        return series
    if number:
        return number
    return stem


def normalize_tax_rate(value: str) -> str:
    value = clean_space(value)
    if not value:
        return ""
    if value.endswith("%"):
        return value
    numeric = parse_decimal(value)
    if numeric == ZERO and re.search(r"0", value):
        return "0%"
    if numeric != ZERO:
        return f"{format_decimal_plain(numeric)}%"
    return value


def extract_tax_rate_from_text(value: str) -> str:
    match = re.search(r"(\d{1,2})\s*%", value)
    return f"{match.group(1)}%" if match else ""


def parse_tax_rate(value: str) -> Optional[Decimal]:
    if not value:
        return None
    cleaned = value.replace("%", "")
    numeric = parse_decimal(cleaned)
    return None if numeric == ZERO and "0" not in cleaned else numeric


def parse_decimal(value: Optional[str]) -> Decimal:
    if value is None:
        return ZERO
    text = clean_space(value)
    if not text:
        return ZERO
    text = text.replace("₫", "").replace("VND", "").replace("đ", "")
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") >= 1 and text.count(".") == 0:
        if text.count(",") == 1 and len(text.split(",")[-1]) <= 2:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return ZERO
    try:
        return Decimal(text)
    except InvalidOperation:
        return ZERO


def format_money(value: Decimal) -> str:
    value = quantize_money(value)
    sign = "-" if value < 0 else ""
    return sign + f"{abs(value):,.0f}"


def quantize_money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)


def format_decimal_plain(value: Decimal) -> str:
    if value == ZERO:
        return ""
    normalized = value.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def decimal_or_zero(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, str):
        return parse_decimal(value)
    return ZERO


def format_date(value: str) -> str:
    value = clean_space(value)
    if not value:
        return ""
    value = value.replace("T00:00:00", "")
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(value, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
    if match:
        yyyy, mm, dd = match.groups()
        return f"{dd}/{mm}/{yyyy}"
    return value


def extract_by_regex(value: str, pattern: str) -> str:
    match = re.search(pattern, value, flags=re.IGNORECASE)
    return clean_space(match.group(1)) if match else ""


def extract_signing_date_from_signature(value: str) -> str:
    return format_date(extract_by_regex(value, r"SigningTime\s*(\d{4}-\d{2}-\d{2})"))


def clean_tax_code(value: str) -> str:
    return re.sub(r"[^0-9-]", "", value or "")


def normalize_compare(value: str) -> str:
    return clean_space(value).casefold()


def unique_preserve_order(values: Iterable[str]) -> List[str]:
    seen = set()
    result = []
    for value in values:
        key = normalize_compare(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def first_non_empty(*values: str) -> str:
    for value in values:
        if clean_space(value):
            return clean_space(value)
    return ""


def map_invoice_nature(value: str) -> str:
    raw = clean_space(value)
    mapping = {
        "1": "Original",
        "2": "Replacement",
        "3": "Adjustment",
        "4": "Replaced",
        "5": "Adjusted",
        "6": "Cancelled",
    }
    return mapping.get(raw, raw)


def write_summary(path: Path, summary: Dict[str, object]) -> None:
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def write_error_csv(path: Path, rows: Sequence[Dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["FileName", "Error"])
        writer.writeheader()
        writer.writerows(rows)


def cast_list(value: object) -> List:
    return value if isinstance(value, list) else []


if __name__ == "__main__":
    main()
