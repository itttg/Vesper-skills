#!/usr/bin/env python3
"""
GFB Billing → SAP Import Converter
Chuyển đổi báo cáo cước Grab For Business sang file import SAP B1

Mỗi dòng GFB tạo 2 dòng SAP:
  - Dòng chi phí: Debit TK 64281001 = PRE_VAT_DELIVERY_FEE + PRE_VAT_SERVICE_FEE
  - Dòng thuế: Debit TK 13311001 = VAT_VALUE_DELIVERY_FEE + VAT_VALUE_SERVICE_FEE
Cuối cùng 1 dòng Credit cho nhà cung cấp = tổng tất cả Debit

Cách dùng:
  python convert_gfb_to_sap.py <input_gfb.xlsx> <output_sap.xlsx> [--posting-date dd/mm/yyyy]
"""

import sys
import argparse
import openpyxl
from datetime import datetime, date
from collections import Counter
import calendar


# ─── Thông tin cố định ──────────────────────────────────────────────────────
EXPENSE_GL      = 64281001
EXPENSE_NAME    = "Chi phí bằng tiền khác"
VAT_GL          = 13311001
VAT_NAME        = "Thuế GTGT được khấu trừ của hàng hóa, dịch vụ"
VENDOR_CODE     = "V00000070"
VENDOR_NAME     = "CÔNG TY TNHH GRAB"
VENDOR_MST      = "312650437"
VENDOR_ADDRESS  = "268 Tô Hiến Thành, Thành phố Hồ Chí Minh, Quận 10"
VENDOR_CONTROL  = 33111001
VENDOR_OFFSET   = 64281001
DISTR_RULE      = "17020101;M999998;M02;ADM;M0100000"
TAX_GROUP       = "PVN5"
BRANCH          = "LEGACY"
PROJECT         = "M02"
PAYMENT_BLOCK   = "N"
TINH_TRANG      = "Kê khai"

# Thứ tự cột SAP (40 cột, A-AN) - khớp chính xác với template
SAP_COLUMNS = [
    'G/L Acct/BP Code',        # A
    'G/L Acct/BP Name',        # B
    'Control Acct',            # C
    'Debit',                   # D
    'Credit',                  # E
    'Distr. Rule',             # F
    'Primary Form Item',       # G
    'CFWId',                   # H
    'Bank Account',            # I
    'Remarks',                 # J
    'Offset Account',          # K
    'Ref. 2',                  # L
    'Due Date',                # M
    'Posting Date',            # N
    'Document Date',           # O
    'Project/Khế ước',         # P
    'Tax Group',               # Q
    'Federal Tax ID',          # R
    'Tax Amount',              # S
    'Gross Value',             # T
    'Base Amount',             # U
    'Payment Block',           # V
    'Block Reason',            # W
    'Branch',                  # X
    'Số HĐ',                  # Y
    'Seri HĐ',                # Z
    'InvType',                 # AA
    'Tình trạng kê khai',     # AB
    'Diễn giải HĐKM',         # AC
    'Nhãn tính C.Nợ',         # AD
    'Mẫu số HĐ',             # AE
    'AdjTran',                 # AF
    'Mã đối tác',             # AG
    'Tên đối tác',            # AH
    'Địa chỉ',               # AI
    'MST',                    # AJ
    'Diễn giải',              # AK
    'RemarksJE',              # AL
    'BP Bank Account',         # AM
    'Share Holder No',         # AN
]


def to_num(v):
    """Chuyển giá trị sang số, trả về 0 nếu không hợp lệ"""
    try:
        return float(v) if v else 0
    except (ValueError, TypeError):
        return 0


def format_amount(amount):
    """Format số tiền cho SAP import dưới dạng số, không thêm hậu tố VND"""
    amt = round(amount)
    if amt == 0:
        return None
    return amt


def format_date_sap(d):
    """Format ngày thành dd.mm.yy cho SAP"""
    if isinstance(d, (date, datetime)):
        return d.strftime('%d.%m.%y')
    if isinstance(d, str):
        try:
            dt = datetime.strptime(str(d)[:10], '%Y-%m-%d')
            return dt.strftime('%d.%m.%y')
        except ValueError:
            return str(d)
    return str(d)


def get_last_day_of_next_month(year, month):
    """Lấy ngày cuối tháng kế tiếp"""
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    last_day = calendar.monthrange(next_year, next_month)[1]
    return date(next_year, next_month, last_day)


def read_gfb_data(filepath):
    """Đọc file GFB Billing Report, trả về list of dict"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            row_dict = dict(zip(headers, row))
            # Lọc: AMOUNT > 0 VÀ phải có COMPANY_NAME (loại dòng tổng cuối file)
            if to_num(row_dict.get('AMOUNT')) > 0 and row_dict.get('COMPANY_NAME'):
                rows.append(row_dict)
    return rows


def detect_billing_month(rows):
    """Xác định tháng chiếm đa số từ VAT_INVOICE_DATE"""
    months = []
    for r in rows:
        d = r.get('VAT_INVOICE_DATE')
        if d:
            months.append(str(d)[:7])  # 'YYYY-MM'
    if not months:
        today = date.today()
        return today.year, today.month
    most_common = Counter(months).most_common(1)[0][0]
    year, month = map(int, most_common.split('-'))
    return year, month


def strip_serial_prefix(serial):
    """Bỏ ký tự '1' đầu tiên của VAT_INVOICE_SERIAL (1C26MGA → C26MGA)"""
    if serial and str(serial).startswith('1'):
        return str(serial)[1:]
    return str(serial) if serial else None


def build_sap_rows(gfb_rows, posting_date, billing_month_str):
    """
    Tạo dữ liệu SAP: mỗi GFB row → 2 SAP rows (chi phí + thuế)
    + 1 dòng credit cuối cùng cho nhà cung cấp
    """
    # Sắp xếp theo VAT_INVOICE_DATE rồi INVOICE_NUMBER
    sorted_rows = sorted(gfb_rows, key=lambda r: (
        str(r.get('VAT_INVOICE_DATE', '')),
        int(r.get('INVOICE_NUMBER', 0)) if r.get('INVOICE_NUMBER') else 0
    ))

    posting_str = format_date_sap(posting_date)
    remarks = f"Chi phí Grab tháng {billing_month_str}_Grab"
    total_debit = 0
    sap_rows = []

    for gfb_row in sorted_rows:
        # Tính toán số tiền
        pre_vat = round(
            to_num(gfb_row.get('PRE_VAT_DELIVERY_FEE')) +
            to_num(gfb_row.get('PRE_VAT_SERVICE_FEE'))
        )
        vat_amt = round(
            to_num(gfb_row.get('VAT_VALUE_DELIVERY_FEE')) +
            to_num(gfb_row.get('VAT_VALUE_SERVICE_FEE'))
        )

        invoice_num = gfb_row.get('INVOICE_NUMBER')
        invoice_date = gfb_row.get('VAT_INVOICE_DATE')
        invoice_serial = strip_serial_prefix(gfb_row.get('VAT_INVOICE_SERIAL'))
        doc_date_str = format_date_sap(invoice_date) if invoice_date else posting_str

        total_debit += pre_vat + vat_amt

        # ── Dòng 1: Chi phí (Debit TK 64281001) ──
        row_exp = {col: None for col in SAP_COLUMNS}
        row_exp.update({
            'G/L Acct/BP Code':     EXPENSE_GL,
            'G/L Acct/BP Name':     EXPENSE_NAME,
            'Control Acct':         EXPENSE_GL,
            'Debit':                format_amount(pre_vat),
            'Distr. Rule':          DISTR_RULE,
            'Remarks':              remarks,
            'Offset Account':       VENDOR_CODE,
            'Due Date':             posting_str,
            'Posting Date':         posting_str,
            'Document Date':        doc_date_str,
            'Project/Khế ước':      PROJECT,
            'Payment Block':        PAYMENT_BLOCK,
            'Branch':               BRANCH,
            'Số HĐ':               invoice_num,
            'Tình trạng kê khai':   TINH_TRANG,
            'Mã đối tác':          VENDOR_CODE,
            'Tên đối tác':         VENDOR_NAME,
            'Địa chỉ':            VENDOR_ADDRESS,
            'MST':                 VENDOR_MST,
            'Diễn giải':          remarks,
            'RemarksJE':           remarks,
        })
        sap_rows.append(row_exp)

        # ── Dòng 2: Thuế GTGT (Debit TK 13311001) ──
        row_vat = {col: None for col in SAP_COLUMNS}
        row_vat.update({
            'G/L Acct/BP Code':     VAT_GL,
            'G/L Acct/BP Name':     VAT_NAME,
            'Control Acct':         VAT_GL,
            'Debit':                format_amount(vat_amt),
            'Remarks':              remarks,
            'Offset Account':       VENDOR_CODE,
            'Due Date':             posting_str,
            'Posting Date':         posting_str,
            'Document Date':        doc_date_str,
            'Project/Khế ước':      PROJECT,
            'Tax Group':            TAX_GROUP,
            'Base Amount':          format_amount(pre_vat),
            'Payment Block':        PAYMENT_BLOCK,
            'Branch':               BRANCH,
            'Số HĐ':               invoice_num,
            'Seri HĐ':             invoice_serial,
            'Tình trạng kê khai':   TINH_TRANG,
            'Mã đối tác':          VENDOR_CODE,
            'Tên đối tác':         VENDOR_NAME,
            'Địa chỉ':            VENDOR_ADDRESS,
            'MST':                 VENDOR_MST,
            'Diễn giải':          remarks,
            'RemarksJE':           remarks,
        })
        sap_rows.append(row_vat)

    # ── Dòng cuối: Credit cho nhà cung cấp ──
    row_credit = {col: None for col in SAP_COLUMNS}
    row_credit.update({
        'G/L Acct/BP Code':     VENDOR_CODE,
        'G/L Acct/BP Name':     VENDOR_NAME,
        'Control Acct':         VENDOR_CONTROL,
        'Credit':               format_amount(total_debit),
        'Remarks':              remarks,
        'Offset Account':       VENDOR_OFFSET,
        'Due Date':             posting_str,
        'Posting Date':         posting_str,
        'Document Date':        posting_str,
        'Project/Khế ước':      PROJECT,
        'Federal Tax ID':       VENDOR_MST,
        'Payment Block':        PAYMENT_BLOCK,
        'Branch':               BRANCH,
        'Tình trạng kê khai':   TINH_TRANG,
        'RemarksJE':           remarks,
    })
    sap_rows.append(row_credit)

    return sap_rows, total_debit


def write_sap_excel(sap_rows, output_path):
    """Ghi file SAP Import Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    for ci, col in enumerate(SAP_COLUMNS, 1):
        ws.cell(row=1, column=ci, value=col)

    # Data rows
    for ri, row_data in enumerate(sap_rows, 2):
        for ci, col in enumerate(SAP_COLUMNS, 1):
            val = row_data.get(col)
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)

        # Cột AJ = MST, luôn xuất dạng text để Excel không hiểu là number
        mst_cell = ws.cell(row=ri, column=36)
        if mst_cell.value is not None:
            mst_cell.number_format = '@'
            mst_cell.value = str(mst_cell.value)

    wb.save(output_path)


def print_summary(gfb_rows, sap_rows, total_debit, posting_date, billing_month_str, output_path):
    """In bảng tóm tắt kết quả"""
    dept_stats = {}
    for r in gfb_rows:
        dept = (r.get('GROUP_NAME') or 'Unknown').strip()
        if dept not in dept_stats:
            dept_stats[dept] = {'trips': 0, 'amount': 0}
        dept_stats[dept]['trips'] += 1
        dept_stats[dept]['amount'] += to_num(r.get('AMOUNT'))

    print(f"\n{'─'*60}")
    print(f"  KẾT QUẢ CHUYỂN ĐỔI GFB → SAP IMPORT")
    print(f"{'─'*60}")
    print(f"  Tháng dữ liệu    : {billing_month_str}")
    print(f"  Posting Date      : {format_date_sap(posting_date)}")
    print(f"  Số chuyến GFB     : {len(gfb_rows)}")
    print(f"  Số dòng SAP       : {len(sap_rows)} ({len(gfb_rows)}x2 + 1 credit)")
    print(f"  Tổng Debit        : {total_debit:,.0f} VND")
    print(f"  File output       : {output_path}")
    print(f"{'─'*60}")
    print(f"\n  Thống kê theo phòng ban:")
    print(f"  {'Phòng ban':<12} {'Chuyến':>6}  {'Tổng tiền':>15}")
    print(f"  {'─'*40}")
    for dept in sorted(dept_stats.keys()):
        d = dept_stats[dept]
        print(f"  {dept:<12} {d['trips']:>6}  {d['amount']:>15,.0f}")
    print(f"  {'─'*40}")
    total_trips = sum(d['trips'] for d in dept_stats.values())
    total_amt = sum(d['amount'] for d in dept_stats.values())
    print(f"  {'TỔNG':<12} {total_trips:>6}  {total_amt:>15,.0f}")
    print()


def main():
    parser = argparse.ArgumentParser(description='GFB Billing → SAP Import')
    parser.add_argument('input', help='File GFB Billing Report (.xlsx)')
    parser.add_argument('output', help='File SAP Import output (.xlsx)')
    parser.add_argument('--posting-date', default=None,
                        help='Ngày hạch toán dd/mm/yyyy (mặc định: cuối tháng kế tiếp)')
    args = parser.parse_args()

    print(f"Đọc file: {args.input}")
    gfb_rows = read_gfb_data(args.input)
    print(f"  → {len(gfb_rows)} chuyến đi hợp lệ (AMOUNT > 0)")

    year, month = detect_billing_month(gfb_rows)
    billing_month_str = f"{month:02d}.{year}"

    if args.posting_date:
        posting_date = datetime.strptime(args.posting_date, '%d/%m/%Y').date()
    else:
        posting_date = get_last_day_of_next_month(year, month)

    print(f"  Tháng dữ liệu: {billing_month_str}")
    print(f"  Posting Date: {format_date_sap(posting_date)}")

    sap_rows, total_debit = build_sap_rows(gfb_rows, posting_date, billing_month_str)
    write_sap_excel(sap_rows, args.output)
    print_summary(gfb_rows, sap_rows, total_debit, posting_date, billing_month_str, args.output)

    # Validation
    gfb_total = sum(to_num(r.get('AMOUNT')) for r in gfb_rows)
    print(f"  VALIDATION:")
    print(f"  Tổng AMOUNT GFB   : {gfb_total:,.0f} VND")
    print(f"  Tổng Debit SAP    : {total_debit:,.0f} VND")
    if abs(gfb_total - total_debit) < 1:
        print(f"  ✓ KHỚP!")
    else:
        diff = gfb_total - total_debit
        print(f"  ✗ CHÊNH LỆCH: {diff:,.0f} VND (có thể do NON_VAT_VALUE)")

    return 0


if __name__ == '__main__':
    sys.exit(main())
